from collections import defaultdict
from typing import Dict, Tuple, List
from openpyxl.worksheet.worksheet import Worksheet
from .file_reader_response import FileReaderResponse
from .project_data import FileReaderProjectData
from spellchecker import SpellChecker
from openpyxl import load_workbook
import re


def tuple_to_str(tup: Tuple[int, int]) -> str:
    ntl = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    resp = ntl[tup[1] + 1]
    resp += str(tup[0])
    return resp


SPELL = SpellChecker()

# ADD TERMS TO DICTIONARY THAT YOU DON'T WANT SPELL-CHECKED
TECH_TERMS = {"MS1", "MS2", "AGC", "RF", "LCM"}


def spell_check(word: str):
    if not word:
        return None

    # Keep only letters, numbers, and spaces
    cleaned = re.sub(r'[^A-Za-z0-9 ]+', ' ', word)
    words = cleaned.split()

    corrected_words = []

    for w in words:
        # Skip empty strings, numbers, and known technical terms
        if not w or w.isnumeric() or w in TECH_TERMS:
            continue

        cor = SPELL.correction(w)
        if cor != w:
            corrected_words.append(cor)

    if not corrected_words:
        return None

    return " ".join(corrected_words)


def add_to_labels(
    labels: Dict[str, Tuple[str, Tuple[int, int]]],
    ran: Tuple[int, int],
    category: str,
    ws: Worksheet
):
    start, end = ran
    for i in range(start, end + 1):
        cell = ws.cell(i, 2)
        if cell is None:
            continue
        labels[cell.value] = (category, (i, 2))


def get_category_label_value(
    labels: Dict[str, Tuple[str, Tuple[int, int]]],
    group_col: int,
    ws: Worksheet,
    typo_correction_location: Dict[str, Dict[str, List[str]]]
) -> Dict[str, Dict[str, str]]:
    """
    labels: {label: (category, (row, col))}
    typo_correction_location: {Possible Typo: {Suggested Correction: [locations found]}}
    Returns: {category: {label: value}}
    """
    cat_lab_val: Dict[str, Dict[str, str]] = defaultdict(dict)

    for label, (category, location) in labels.items():
        row = location[0]
        val = ws.cell(row, group_col).value

        cor = spell_check(val)
        if cor is not None:
            typo_correction_location[val][cor].append(tuple_to_str((row, group_col)))

        cat_lab_val[category][label] = val

    return cat_lab_val


# {Group: {Category: {Label: Value}}}


def get_independent_variables(
    group_category_label_value: Dict[str, Dict[str, Dict[str, str]]]
) -> Dict[str, List[str]]:
    # {Label: set of all values across groups}
    labels = defaultdict(set)

    for group in group_category_label_value:
        for cat in group_category_label_value[group]:
            for label, value in group_category_label_value[group][cat].items():
                labels[label].add(value)

    # Only keep labels with more than one unique value
    ind_vars = defaultdict(list)
    for label, values in labels.items():
        if len(values) > 1:
            ind_vars[label].extend(values)

    return ind_vars



class FileReader:
    def __init__(self):
        self.sample_id_range = (10, 24)
        self.sample_prep_range = (26, 43)
        self.lc_param_range = (45, 66)
        self.ms_param_range = (68, 89)

    def get_file_reader_response(
        self,
        project_data: FileReaderProjectData,
        file_path: str
    ) -> FileReaderResponse:
        success = True
        message = "Data Read Successfully!"

        # {Data_Tag: [Values]}
        independent_variables: Dict[str, List[str]] = defaultdict(list)
        # {Group: {Category: {Label: Value}}}
        data: Dict[str, Dict[str, Dict[str, str]]] = defaultdict(lambda: defaultdict(dict))
        # {Possible Typo: {Suggested Correction: [locations found]}}
        possible_typos: Dict[str, Dict[str, List[str]]] = defaultdict(lambda: defaultdict(list))

        resp = FileReaderResponse(
            success=success,
            message=message,
            project_data=project_data,
            independent_variables=independent_variables,
            data=data,
            possible_typos=possible_typos
        )

        try:
            wb = load_workbook(file_path, read_only=True)
            ws = wb["DataEntry"]
        except Exception as e:
            resp.success = False
            resp.message = f"Failed To Read This File Due To Exception: {e}"
            return resp

        expected_groups = project_data.get_groups()

        # Labels == {label: (category, (row, col))}
        labels: Dict[str, Tuple[str, Tuple[int, int]]] = {}

        for ran, category in [
            (self.sample_id_range, "Sample ID"),
            (self.sample_prep_range, "Sample Prep"),
            (self.lc_param_range, "LC Param"),
            (self.ms_param_range, "MS Param")
        ]:
            add_to_labels(labels, ran, category, ws)

        # {Group: {Category: {Label: Value}}}
        group_category_label_value: Dict[str, Dict[str, Dict[str, str]]] = defaultdict(lambda: defaultdict(dict))

        # {Possible Typo: {Suggested Correction: [locations found]}}
        typo_correction_location: Dict[str, Dict[str, List[str]]] = defaultdict(lambda: defaultdict(list))

        for group_index, expected in enumerate(expected_groups):
            group_col = group_index + 3
            read_group = ws.cell(8, group_col).value

            if expected != read_group:
                resp.success = False
                resp.message = (
                    f"Uploaded Sheet Groups Didn't Correspond "
                    f"To Expected Groups (at[{8},{group_col}])\n"
                    f"EXPECTED: {expected} != GOT: {read_group}"
                )
                return resp

            group_category_label_value[read_group] = get_category_label_value(
                labels, group_col, ws, typo_correction_location
            )

        # {Data_Tag: [Values]}
        ind_vars = get_independent_variables(group_category_label_value)

        # Attach processed data to response
        resp.data = group_category_label_value
        resp.possible_typos = typo_correction_location
        resp.independent_variables = ind_vars

        return resp
