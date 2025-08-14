import copy
import datetime
import os
import shutil

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Protection, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

from core.utils.excel.project_data import FileReaderProjectData


class ExcelFileGenerator:

    def __init__(self):
        base_dir = os.path.dirname(os.path.abspath(__file__))
        self.input_directory_path = os.path.join(base_dir, "excel_templates")
        self.output_directory_path = os.path.join(base_dir, "output_files")
        self.sample_id_range = [(10, 19)]
        self.sample_prep_range = [(26, 38)]
        self.lc_param_range = [(45, 61)]
        self.ms_param_range = [(68, 84)]
        self.custom_ranges = [(20, 24), (39, 43), (62, 66), (85, 89)]
        self.black_ranges = [(9, 9), (25, 25), (44, 44), (67, 67)]

    def make_a_copy(self, input_file: str, name=None) -> str:
        date = datetime.datetime.now().replace(microsecond=0).isoformat().replace(":", "-")

        old_file_path = os.path.join(self.input_directory_path, input_file)
        file_type = old_file_path.split(".")[-1]
        new_file_path = (
                            os.path.join(self.output_directory_path, date)
                            if name is None
                            else os.path.join(self.output_directory_path, name + "-" + date)
                        ) + "." + file_type

        shutil.copy2(old_file_path, new_file_path)

        return new_file_path

    def make_new_file_from_template_with_openpyxl(self, project_data: FileReaderProjectData, output_name=None) -> str:
        INPUT_FILE = "metadataTemplate6.xlsm"
        file = self.make_a_copy(INPUT_FILE, output_name)
        wb = load_workbook(file, keep_vba=True)
        ws = wb["DataEntry"]

        self.add_dropdowns(ws)

        self.add_project_data(ws, project_data)

        wb.save(file)
        return file

    def add_dropdowns(self, ws) -> None:
        dropdowns = [
            (self.sample_id_range[0][0], 1),
            (self.sample_prep_range[0][0], 18),
            (self.lc_param_range[0][0], 38),
            (self.ms_param_range[0][0], 62),
        ]

        for start_row, source_row in dropdowns:
            dv = DataValidation(
                type="list",
                formula1=f"=Source!$C${source_row}:$N${source_row}",
                allow_blank=True,
            )
            ws.add_data_validation(dv)
            dv.add(f"A{start_row + 1}")

    def add_project_data(self, ws, project_data: FileReaderProjectData) -> None:
        """
        :param ws: the current worksheet
        :param project_data: the project data, including name, description, and groups []
        :return: Nothing

        1. Add Title
        2. For Group in Groups
            a. Write group title
            b. For section in sections in ranges []
                i. unlock it
                ii. color it the right color
        3. protect sheet?
        """

        # Set the project Name
        ws["A1"].value = project_data.get_name()

        # Get the groups from Project Data
        groups = project_data.get_groups()

        section_ranges = [self.sample_id_range, self.sample_prep_range, self.black_ranges,
                          self.lc_param_range, self.ms_param_range, self.custom_ranges]

        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        def make_group_title(c):
            group_name_cell = ws.cell(row=8, column=c)
            group_name_cell.value = groups[c - 3]
            group_name_cell.font = Font(bold=True, size=18)
            group_name_cell.fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

        for group in range(len(groups)):
            col = 3 + group
            make_group_title(col)
            for section_range in section_ranges:
                cur_fill = ws.cell(section_range[0][0], 3).fill
                for chunk in section_range:
                    start, end = chunk
                    for row in range(start, end + 1):
                        cell = ws.cell(row=row, column=col)
                        cell.fill = copy.copy(cur_fill)
                        cell.border = thin_border
                        cell.protection = Protection(locked=False)
